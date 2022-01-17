#!python
#!/usr/bin/env python

import sys
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.metrics import classification_report,confusion_matrix
from sklearn.metrics import  f1_score
from xgboost import XGBClassifier
import pandas as pd

from sklearn.ensemble import VotingClassifier
from sklearn.compose import ColumnTransformer
from sklearn.preprocessing import FunctionTransformer
from sklearn.pipeline import Pipeline
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis
from sklearn.neighbors import KNeighborsClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.svm import SVC
import openpyxl
import warnings
warnings.filterwarnings("ignore")

def read_dat_dict_file(read_file):

    archaea_id = dict()
    classes = dict()
    sq_len = dict()
    sq_gc = dict()
    nc = dict()
    nc_ir0 = dict()
    nc_ir1 = dict()
    nc_ir2 = dict()

    data_tmp = np.load(read_file, allow_pickle=True).item()
    for i, archaea in enumerate(data_tmp['archaea_id']):
        
        classes[archaea] = data_tmp['classes'][i]
        sq_len[archaea] = data_tmp['sq_len'][i]
        sq_gc[archaea] = data_tmp['sq_gc'][i]
        nc[archaea] = data_tmp['nc'][i]
        nc_ir0[archaea] = data_tmp['nc_ir0'][i]
        nc_ir1[archaea] = data_tmp['nc_ir1'][i]
        nc_ir2[archaea] = data_tmp['nc_ir2'][i]

        return classes, sq_len, sq_gc, nc, nc_ir0, nc_ir1, nc_ir2

def read_data_file(read_data_file,read_label_file):
    labels_tmp = np.load(read_label_file, allow_pickle=True)
    data_tmp = np.load(read_data_file, allow_pickle=True)
    return data_tmp, labels_tmp

def xgboost_classification(data, labels, taxa, variables, iterations):
        print("Features: "+variables)
        AC=[]
        F1=[]
        p_ac_list=[]
        print("Using XGBoost for "+taxa+"...")
        print("Number of samples: ", np.shape(data)[0])
        print("Number of Labels: ", np.shape(np.unique(labels))[0])
        for a in range(iterations):
            X_train, X_test, y_train, y_test = train_test_split(data, labels, test_size=0.20, stratify=labels, random_state=a)
            model = XGBClassifier(max_depth=6,learning_rate=0.1, n_estimators=250,eval_metric='mlogloss')
            model.fit(X_train, y_train)
            y_pred = model.predict(X_test)
            predictions = [round(value) for value in y_pred]
            p_ac_list.append(determine_class_accuracy(y_test,predictions))
            
            AC.append(accuracy_score(y_test, predictions))
            F1.append(f1_score(y_test, y_pred, average='weighted'))

        avg_ac = sum(AC)/len(AC)
        avg_f1 = sum(F1)/len(F1) 
        
        print("Average Accuracy of : %.2f%%" % (avg_ac * 100.0),"Max: %.2f%%" % (max(AC)* 100.0), "Min: %.2f%%" % (min(AC)* 100.0) )
        print("F1 score of : %.2f%%" % (avg_f1),"Max: %.2f%%" % (max(F1)),"Min: %.2f%%" % (min(F1)) )
        print("--------------------------")
        index_max = AC.index(max(AC))
        p_ac=p_ac_list[index_max]
        hit_probability=determine_hit_probability(p_ac,labels)

        ac_list = [taxa, np.shape(np.unique(labels))[0],np.shape(data)[0], avg_ac* 100.0 ]
        f1_s = [taxa, np.shape(np.unique(labels))[0],np.shape(data)[0], avg_f1 ]
        return  ac_list, f1_s, hit_probability

def selector(X):
    return X

import numpy as np
from sklearn.preprocessing import LabelEncoder

def fit_multiple_estimators(classifiers, X_list, y, sample_weights = None):

    # Convert the labels `y` using LabelEncoder, because the predict method is using index-based pointers
    # which will be converted back to original data later.
    le_ = LabelEncoder()
    le_.fit(y)
    transformed_y = le_.transform(y)

    # Fit all estimators with their respective feature arrays
    estimators_ = [clf.fit(X, y) if sample_weights is None else clf.fit(X, y, sample_weights) for clf, X in zip([clf for _, clf in classifiers], X_list)]

    return estimators_, le_


def predict_from_multiple_estimator(estimators, label_encoder, X_list, weights = None):

    # Predict 'soft' voting with probabilities

    pred1 = np.asarray([clf.predict_proba(X) for clf, X in zip(estimators, X_list)])
    pred2 = np.average(pred1, axis=0, weights=weights)
    pred = np.argmax(pred2, axis=1)

    # Convert integer predictions to original labels:
    return label_encoder.inverse_transform(pred)

def xgboost_voting_classification(data, labels, taxa, variables, iterations):
        print("Features: "+variables)
        AC=[]
        F1=[]
        p_ac_list=[]
        print("Using XGBoost for "+taxa+"...")
        print("Number of samples: ", np.shape(data)[0])
        print("Number of Labels: ", np.shape(np.unique(labels))[0])
        for a in range(iterations):
            X_train, X_test, y_train, y_test = train_test_split(data, labels, test_size=0.20, stratify=labels, random_state=a)            
            X_train1, X_train2 = X_train[:,:3], X_train[:,3:]
            X_test1, X_test2 = X_test[:,:3], X_test[:,3:]
            X_train_list = [X_train1, X_train2]
            X_test_list = [X_test1, X_test2]
            classifiers = [('xgb1',  XGBClassifier(max_depth=6,learning_rate=0.1, n_estimators=250,eval_metric='mlogloss')),
                            ('xgb2', XGBClassifier(max_depth=6,learning_rate=0.1, n_estimators=150,eval_metric='mlogloss'))]
            fitted_estimators, label_encoder = fit_multiple_estimators(classifiers, X_train_list, y_train)
            y_pred = predict_from_multiple_estimator(fitted_estimators, label_encoder, X_test_list,[2,1])
            predictions = [round(value) for value in y_pred]
            p_ac_list.append(determine_class_accuracy(y_test,predictions))
            AC.append(accuracy_score(y_test, predictions))
            F1.append(f1_score(y_test, y_pred, average='weighted'))

        avg_ac = sum(AC)/len(AC)
        avg_f1 = sum(F1)/len(F1) 
        
        print("Average Accuracy of : %.2f%%" % (avg_ac * 100.0),"Max: %.2f%%" % (max(AC)* 100.0), "Min: %.2f%%" % (min(AC)* 100.0) )
        print("F1 score of : %.2f%%" % (avg_f1),"Max: %.2f%%" % (max(F1)),"Min: %.2f%%" % (min(F1)) )
        print("--------------------------")
        index_max = AC.index(max(AC))
        p_ac=p_ac_list[index_max]
        hit_probability=determine_hit_probability(p_ac,labels)

        ac_list = [taxa, np.shape(np.unique(labels))[0],np.shape(data)[0], avg_ac* 100.0 ]
        f1_s = [taxa, np.shape(np.unique(labels))[0],np.shape(data)[0], avg_f1 ]
        return  ac_list, f1_s, hit_probability



def other_classifications(data, labels, taxa, variables, iterations):
        AC_LDA=[]
        AC_GNB=[]
        AC_SVC=[]
        AC_KNN=[]
        F1_LDA=[]
        F1_GNB=[]
        F1_SVC=[]
        F1_KNN=[]

        print("Classification of "+taxa+"...")
        print("Features: "+variables)
        print("Number of samples: ", np.shape(data)[0])
        print("Number of Labels: ", np.shape(np.unique(labels))[0])
        
        for a in range(iterations):
            X_train, X_test, y_train, y_test = train_test_split(data, labels, test_size=0.4, stratify=labels, random_state=a)
            
            LDA_model = LinearDiscriminantAnalysis()
            GaussianNB_model = GaussianNB()
            SVC_model = SVC()
            KNN_model = KNeighborsClassifier(n_neighbors=np.shape(np.unique(labels))[0])
            
            LDA_model.fit(X_train, y_train)
            GaussianNB_model.fit(X_train, y_train)
            SVC_model.fit(X_train, y_train)
            KNN_model.fit(X_train, y_train)

            LDA_prediction = LDA_model.predict(X_test)
            GaussianNB_prediction = GaussianNB_model.predict(X_test)
            SVC_prediction = SVC_model.predict(X_test)
            KNN_prediction = KNN_model.predict(X_test)
    
            AC_LDA.append(accuracy_score(y_test, LDA_prediction))
            AC_GNB.append(accuracy_score(y_test, GaussianNB_prediction))
            AC_SVC.append(accuracy_score(y_test, SVC_prediction))
            AC_KNN.append(accuracy_score(y_test, KNN_prediction))

            F1_LDA.append(f1_score(y_test, LDA_prediction, average='weighted'))
            F1_GNB.append(f1_score(y_test, GaussianNB_prediction, average='weighted'))
            F1_SVC.append(f1_score(y_test, SVC_prediction, average='weighted'))
            F1_KNN.append(f1_score(y_test, KNN_prediction, average='weighted'))
        #Average Accuracy
        avg_ac_lda = sum(AC_LDA)/len(AC_LDA)
        avg_ac_gnb = sum(AC_GNB)/len(AC_GNB)
        avg_ac_svc = sum(AC_SVC)/len(AC_SVC)
        avg_ac_knn = sum(AC_KNN)/len(AC_KNN)
        
        #Average F1-score
        avg_f1_lda = sum(F1_LDA)/len(F1_LDA) 
        avg_f1_gnb = sum(F1_GNB)/len(F1_GNB) 
        avg_f1_svc = sum(F1_SVC)/len(F1_SVC) 
        avg_f1_knn = sum(F1_KNN)/len(F1_KNN) 

        print("LDA Average Accuracy of : %.2f%%" % (avg_ac_lda * 100.0),"Max: %.2f%%" % (max(AC_LDA)* 100.0), "Min: %.2f%%" % (min(AC_LDA)* 100.0) )
        print("LDA F1 score of : %.2f%%" % (avg_f1_lda),"Max: %.2f%%" % (max(F1_LDA)),"Min: %.2f%%" % (min(F1_LDA)) )
        print("GNB Average Accuracy of : %.2f%%" % (avg_ac_gnb * 100.0),"Max: %.2f%%" % (max(AC_GNB)* 100.0), "Min: %.2f%%" % (min(AC_GNB)* 100.0) )
        print("GNB F1 score of : %.2f%%" % (avg_f1_gnb),"Max: %.2f%%" % (max(F1_GNB)),"Min: %.2f%%" % (min(F1_GNB)) )
        print("SVC Average Accuracy of : %.2f%%" % (avg_ac_svc * 100.0),"Max: %.2f%%" % (max(AC_SVC)* 100.0), "Min: %.2f%%" % (min(AC_SVC)* 100.0) )
        print("SVC F1 score of : %.2f%%" % (avg_f1_svc),"Max: %.2f%%" % (max(F1_SVC)),"Min: %.2f%%" % (min(F1_SVC)) )
        print("KNN Average Accuracy of : %.2f%%" % (avg_ac_knn * 100.0),"Max: %.2f%%" % (max(AC_KNN)* 100.0), "Min: %.2f%%" % (min(AC_KNN)* 100.0) )
        print("KNN F1 score of : %.2f%%" % (avg_f1_svc),"Max: %.2f%%" % (max(F1_KNN)),"Min: %.2f%%" % (min(F1_KNN)) )
        print("--------------------------")
        

        Acc=[taxa, np.shape(np.unique(labels))[0],np.shape(data)[0], avg_ac_lda* 100.0,avg_f1_gnb* 100.0, avg_f1_svc* 100.0, avg_f1_knn* 100.0]
        f1_s=[taxa, np.shape(np.unique(labels))[0],np.shape(data)[0], avg_f1_lda,avg_f1_gnb, avg_f1_svc, avg_f1_knn]
        return Acc, f1_s

def determine_class_accuracy(y_true,y_pred):
    matrix = confusion_matrix(y_true, y_pred)
    matrix_list= matrix.diagonal()
    divisor=matrix.sum(axis=1).tolist()
    division = [a / b if a>0 and b>0 else 0 for a, b in zip(matrix_list, divisor)]
    return division

def determine_random_hit_percentage(labels):
    occurance=[[x,labels.tolist().count(x)] for x in set(labels.tolist())]
    p_occurance=[x[1]/len(labels) for x in occurance]
    number_classes=len(set(labels.tolist()))
    class_accuracy=[1/number_classes]*number_classes
    return hit_percentage(p_occurance,class_accuracy)

def determine_hit_probability(p_ac,labels):
    occurance=[[x,labels.tolist().count(x)] for x in set(labels.tolist())]
    p_occurance=[x[1]/len(labels) for x in occurance]
    number_classes=len(set(labels.tolist()))
    return hit_percentage(p_occurance,p_ac)

def hit_percentage(p_occurance, class_accuracy):
    p_hit=sum([a * b for a, b in zip(p_occurance, class_accuracy)])*100
    return p_hit

def classification(relative_path):
    taxonomy=["Phylum","Class","Order","Family","Genus"]
    iterations=5
    acc_xgboost=[["Classification","N. Classes","Samples","NC","NC_gen+NC_pr","SL+GC+NC(genome)","SL+GC+NC(proteome)","All"]]
    f1_xgboost=[["Classification","N. Classes","Samples","NC","NC_gen+NC_pr","SL+GC+NC(genome)","SL+GC+NC(proteome)","All"]]
    acc_other=[["Classification","N. Classes","Samples","LDA","GNB","SVM","KNN","XGB"]]
    f1_other=[["Classification","N. Classes","Samples","LDA","GNB","SVM","KNN","XGB"]]
    p_hit=[]
    for tx in taxonomy:
        
        x_file="../data/"+tx+"_x_data.npy"
        y_file="../data/"+tx+"_y_data.npy"
        data, labels = read_data_file(x_file,y_file)
        nc_data=data[:,[2]]
        ncs_data=data[:,[2,4]]
        sq_gc_nc_genome=data[:,:3]
        sq_nc_protein=data[:,3:]
        
        
        random_hit_per=determine_random_hit_percentage(labels)
        #xgboost Classification
        ac_1f, f1_1f, _ = xgboost_classification(nc_data, labels, tx, "NC", iterations)
        ac_ncs, f1_ncs, _ = xgboost_classification(ncs_data, labels, tx, "NC_genome+NC_proteome", iterations)
        ac_gen, f1_gen, _ = xgboost_classification(sq_gc_nc_genome, labels, tx, "SL+GC+NC(genome)", iterations)
        ac_pro, f1_pro, _ = xgboost_classification(sq_nc_protein, labels, tx, "SL+NC(proteome)", iterations)
        ac_all, f1_all, class_hit_per = xgboost_voting_classification(data, labels, tx, "All", iterations)
        
        #p_hit Classification
        p_hit.append([tx,random_hit_per,class_hit_per])
        
        #Other Classifiers
        ac_other, f1_l_other = other_classifications(data, labels, tx, "All", iterations)
        
        acc_xgboost.append(ac_1f + [ac_ncs[-1]] + [ac_gen[-1]] + [ac_pro[-1]] + [ac_all[-1]])
        f1_xgboost.append(f1_1f + [f1_ncs[-1]] + [f1_gen[-1]] + [f1_pro[-1]] + [f1_all[-1]])
        acc_other.append(ac_other + [ac_all[-1]])
        f1_other.append(f1_l_other + [f1_all[-1]])
    
    path_p_hit = relative_path + "/" + "p_hit_class.xlsx"
    path_ac_xgb = relative_path + "/" + "Acc_Xgboost.xlsx"
    path_f1_xgb = relative_path + "/" + "F1_Xgboost.xlsx"
    path_ac_all = relative_path + "/" + "Acc_other_class.xlsx"
    path_f1_all = relative_path + "/" + "F1_other_class.xlsx"

    to_csv(path_p_hit,p_hit)
    to_csv(path_ac_xgb, acc_xgboost)
    to_csv(path_f1_xgb, f1_xgboost)
    to_csv(path_ac_all, acc_other)
    to_csv(path_f1_all, f1_other)

def to_csv(path, values_l): 

    wb = openpyxl.Workbook() 
    wb.save(filename = path)
    
    df = pd.DataFrame(values_l)
    book = openpyxl.load_workbook(path)
    writer = pd.ExcelWriter(path, engine="openpyxl",mode='a')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer)
    writer.save()
    wb = openpyxl.load_workbook(path)
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(path)

if __name__ == "__main__":
    xls_path="../xlslist"
    classification(xls_path)
